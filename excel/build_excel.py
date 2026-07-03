import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

df = pd.read_csv("/home/claude/cardio/data/excel_cleaned_cardio_data.csv")

HEADER_FILL = PatternFill("solid", fgColor="C0392B")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
TITLE_FONT = Font(bold=True, color="C0392B", name="Arial", size=14)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------- Sheet 1: Cleaned Data ----------
ws = wb.active
ws.title = "Cleaned_Data"
# A id, B age, C age_group, D gender_label, E bmi, F bmi_category, G ap_hi, H ap_lo,
# I bp_category, J cholesterol_label, K gluc_label, L smoke, M alco, N active,
# O risk_score, P risk_tier, Q cardio
cols = ["id","age","age_group","gender_label","bmi","bmi_category","ap_hi","ap_lo","bp_category",
        "cholesterol_label","gluc_label","smoke","alco","active","risk_score","risk_tier","cardio"]
ws.append(cols)
for cell in ws[1]:
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = Alignment(horizontal="center")
for _, row in df[cols].head(2000).iterrows():
    ws.append(list(row))
for i, c in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(i)].width = 14
ws.freeze_panes = "A2"
n = 2000

# ---------- Sheet 2: KPI Dashboard ----------
kpi = wb.create_sheet("KPI_Dashboard")
kpi["A1"] = "Cardiovascular Risk Analysis - KPI Dashboard"
kpi["A1"].font = TITLE_FONT
kpi.merge_cells("A1:D1")

kpi["A3"] = "Metric"; kpi["B3"] = "Value"
for c in ("A3","B3"):
    kpi[c].fill = HEADER_FILL; kpi[c].font = HEADER_FONT

kpi["A4"] = "Total Patients"
kpi["B4"] = f"=COUNTA(Cleaned_Data!A2:A{n+1})"
kpi["A5"] = "Positive Cardio Cases"
kpi["B5"] = f"=SUM(Cleaned_Data!Q2:Q{n+1})"
kpi["A6"] = "Overall Disease Rate (%)"
kpi["B6"] = "=ROUND(B5/B4*100,2)"
kpi["A7"] = "Average Risk Score"
kpi["B7"] = f"=ROUND(AVERAGE(Cleaned_Data!O2:O{n+1}),1)"
kpi["A8"] = "Average BMI"
kpi["B8"] = f"=ROUND(AVERAGE(Cleaned_Data!E2:E{n+1}),1)"
kpi["A9"] = "Average Systolic BP"
kpi["B9"] = f"=ROUND(AVERAGE(Cleaned_Data!G2:G{n+1}),1)"
kpi["A10"] = "Smokers with Disease (%)"
kpi["B10"] = f'=ROUND(SUMIFS(Cleaned_Data!Q2:Q{n+1},Cleaned_Data!L2:L{n+1},1)/SUMIF(Cleaned_Data!L2:L{n+1},1)*100,2)'
for r in range(4, 11):
    kpi[f"A{r}"].font = BODY_FONT; kpi[f"B{r}"].font = Font(bold=True, name="Arial", color="C0392B")
    kpi[f"A{r}"].border = BORDER; kpi[f"B{r}"].border = BORDER
kpi.column_dimensions["A"].width = 30
kpi.column_dimensions["B"].width = 14

kpi["D3"] = "Risk Tier"; kpi["E3"] = "Patients"; kpi["F3"] = "Disease Rate (%)"
for c in ("D3","E3","F3"):
    kpi[c].fill = HEADER_FILL; kpi[c].font = HEADER_FONT
tiers = ["Low","Medium","High","Critical"]
for i, t in enumerate(tiers, 4):
    kpi[f"D{i}"] = t
    kpi[f"E{i}"] = f'=COUNTIF(Cleaned_Data!P2:P{n+1},D{i})'
    kpi[f"F{i}"] = f'=IF(E{i}=0,0,ROUND(SUMIFS(Cleaned_Data!Q2:Q{n+1},Cleaned_Data!P2:P{n+1},D{i})/E{i}*100,2))'
    for col in "DEF":
        kpi[f"{col}{i}"].border = BORDER; kpi[f"{col}{i}"].font = BODY_FONT
kpi.column_dimensions["D"].width = 14
kpi.column_dimensions["E"].width = 12
kpi.column_dimensions["F"].width = 16

chart = BarChart()
chart.title = "Disease Rate by Risk Tier"
chart.y_axis.title = "Disease Rate (%)"
data = Reference(kpi, min_col=6, min_row=3, max_row=7)
cats = Reference(kpi, min_col=4, min_row=4, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
kpi.add_chart(chart, "D10")

# ---------- Sheet 3: Age Group Pivot ----------
piv = wb.create_sheet("Age_Group_Analysis")
age_groups = sorted(df["age_group"].dropna().unique().tolist(), key=lambda x: df.loc[df.age_group==x, "age"].mean())
piv.append(["Age Group","Patients","Disease Rate (%)","Avg BMI","Avg Systolic BP"])
for cell in piv[1]:
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
for i, g in enumerate(age_groups, 2):
    piv[f"A{i}"] = g
    piv[f"B{i}"] = f'=COUNTIF(Cleaned_Data!C2:C{n+1},A{i})'
    piv[f"C{i}"] = f'=IF(B{i}=0,0,ROUND(AVERAGEIF(Cleaned_Data!C2:C{n+1},A{i},Cleaned_Data!Q2:Q{n+1})*100,2))'
    piv[f"D{i}"] = f'=IF(B{i}=0,0,ROUND(AVERAGEIF(Cleaned_Data!C2:C{n+1},A{i},Cleaned_Data!E2:E{n+1}),1))'
    piv[f"E{i}"] = f'=IF(B{i}=0,0,ROUND(AVERAGEIF(Cleaned_Data!C2:C{n+1},A{i},Cleaned_Data!G2:G{n+1}),1))'
for col, w in zip("ABCDE", [12,10,16,10,16]):
    piv.column_dimensions[col].width = w

chart2 = BarChart()
chart2.title = "Disease Rate by Age Group"
data2 = Reference(piv, min_col=3, min_row=1, max_row=1+len(age_groups))
cats2 = Reference(piv, min_col=1, min_row=2, max_row=1+len(age_groups))
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
piv.add_chart(chart2, "G2")

wb.save("/home/claude/cardio/excel/cardio_project.xlsx")
print("saved")
